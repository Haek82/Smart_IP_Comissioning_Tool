from openpyxl import load_workbook
import json
import sys
import datetime

class Xlsx:
    #This is the excel import / export class

    def __init__(self, filename):

        self.workbook = load_workbook(filename=filename)
        self.sheet = self.workbook.active
        self.speakers = []
        self.importFromExcel()
        self.filename = filename

    def validate_ip(self, s):
        #print(s)
        if s is None:
            return False
        a = s.split('.')
        if len(a) != 4:
            return False
        for x in a:
            if not x.isdigit():
                return False
            i = int(x)
            if i < 0 or i > 255:
                return False
        return True


    def importFromExcel(self):
        for row in self.sheet.iter_rows(min_row=4, min_col=1, max_col=5, values_only=True):
            if row[0] is not None:
                if row[0].find("4420") != -1 or row[0].find("4430") != -1:
                    if self.validate_ip(row[1]) is True and self.validate_ip(row[2]) is True and self.validate_ip(row[3]) is True:
                        speaker = {
                            "barcode" : row[0],
                            "ip": row[1],
                            "mask": row[2],
                            "gw": row[3]
                            }
                        self.speakers.append(speaker)
                    else:
                        print("Ip infromation in excel wrong or missing")
                        sys.exit()
                else:
                    print("Barcode column contains non valid serialnumber")
                    sys.exit()


    def setDate(self):
        x = datetime.datetime.now()
        dateCell = "G1"
        self.sheet[dateCell] = x.strftime("%d-%m-%y %H:%M:%S")
        self.workbook.save(filename=self.filename)

    def setDanteNameAndMac(self, row, name, mac):
        macCell = "E" + str(row+2)
        danteCell = "F" + str(row+2)
        updatedCell = "G" + str(row+2)
        self.sheet[macCell] = mac
        self.sheet[danteCell] = name
        self.sheet[updatedCell] = "X"
        self.workbook.save(filename=self.filename)

    def printAllSpeakersInExcel(self):
        print("Speaker infomation in excel: ")
        for speaker in self.speakers:
            print(speaker)

    def getList(self):
        return self.speakers
